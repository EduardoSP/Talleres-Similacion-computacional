#import xlwt
from datetime import datetime
from openpyxl import Workbook

def write_text_excel(name_file, title, table, number_rows, number_columns, list_text_head):
	row = 4
	style0 = xlwt.easyxf('font: name Times New Roman, colour red, bold on')
	style1 = xlwt.easyxf('',num_format_str='DD-MMM-YY')
	wb = xlwt.Workbook()
	ws = wb.add_sheet(title, cell_overwrite_ok=True)
	ws.write(0, 0, title, style0)
	ws.write(1, 0, datetime.now(), style1)
	for c in range(number_columns): #write head
		ws.write(row, c, list_text_head[c])
	row += 1
	for i in range(number_rows):
		row += 1
		for k in range(number_columns):
			ws.write(row, k, table[i][k])
	name_file = name_file + '.xls'
	wb.save(name_file) #save file
	print("Excel creado")


def write_text_excel_openpyxl(name_file, title, table, number_rows, number_columns, list_text_head):
	list_columns = ["A", "B", "C", "D", "E", "F","G", "H", "I","J", "K", "L", "M", "N"]
	wb = Workbook()
	# grab the active worksheet
	ws = wb.active
	# Data can be assigned directly to cells
	ws["A1"] = title
	# Python types will automatically be converted
	ws['A2'] = datetime.now()
	row = 4
	pos = ''
	for c in range(number_columns): #write head
		pos = str(list_columns[c]) + str(row)
		ws[pos] = list_text_head[c] 
	row += 1
	for i in range(number_rows):
		row += 1
		for k in range(number_columns):
			pos = str(list_columns[k]) + str(row)
			ws[pos] = table[i][k]

	name_file = name_file + ".xlsx"
	# Python types will automatically be converted
	# Save the file
	wb.save(name_file)
	print("Excel creado")
